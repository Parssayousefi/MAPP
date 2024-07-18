import openpyxl
import os
import re

def process_questionnaire(file_path):
    # Read the text file
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    # Split the content into questionnaires
    questionnaires = re.split(r'Start of Block: (.*?)\n', content)[1:]  # Skip the first empty element

    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Questionnaire Items"

    # Add headers
    sheet['A1'] = "Questionnaire Name"
    sheet['B1'] = "Item Number"
    sheet['C1'] = "Content"

    row = 2  # Start from the second row (after headers)

    # Process each questionnaire
    for i in range(0, len(questionnaires), 2):
        questionnaire_name = questionnaires[i].strip()
        questionnaire_content = questionnaires[i+1].split('End of Block:')[0].strip()

        # Split the content into items
        items = re.split(r'\n\n+', questionnaire_content)

        for item in items:
            # Try to extract item number and content
            match = re.match(r'([A-Za-z0-9_\.]+)\s+(.*)', item, re.DOTALL)
            if match:
                item_number = match.group(1)
                item_content = match.group(2).strip()
            else:
                item_number = ""
                item_content = item.strip()

            # Only add the row if either item_number or item_content is not empty
            if item_number or item_content:
                sheet.cell(row=row, column=1, value=questionnaire_name)
                sheet.cell(row=row, column=2, value=item_number)
                sheet.cell(row=row, column=3, value=item_content)
                row += 1

    # Remove empty rows
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3):
        if all(cell.value is None or cell.value == "" for cell in row):
            sheet.delete_rows(row[0].row, 1)

    # Get the directory of the input file
    input_dir = os.path.dirname(file_path)

    # Create the output file path in the same directory as the input file
    output_file = os.path.join(input_dir, "Questionnaire_Items.xlsx")

    # Save the workbook
    workbook.save(output_file)
    print(f"Excel file '{output_file}' has been created.")

# Usage
file_path = r"C:\Users\prsyu\OneDrive\Bidlung\University\M.S. Leiden University\M.S. Neuroscience (Research)\MAPP\Analysis\Github_Clone\MAPP\Questionairs\Scale_Survey.txt"
process_questionnaire(file_path)